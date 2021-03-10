import discord, openpyxl, asyncio, aiohttp, os, sqlite3, json
from discord.utils import get
from discord.ext import commands
from openpyxl.styles import Font
from voltorb import voltorb
from io import BytesIO
from dotenv import load_dotenv

load_dotenv()
from tags import Tags

prefix = '%'
sections = ['CE-A', 'CE-B', 'CE-C', 'CS-A', 'CS-B', 'EC-A', 'EC-B', 'EC-C', 'EE-A', 'EE-B', 'EE-C', 'IT-A', 'IT-B', 'ME-A', 'ME-B', 'ME-C', 'PI-A', 'PI-B']
subsections = ['CE-01', 'CE-02', 'CE-03', 'CE-04', 'CE-05', 'CE-06', 'CE-07', 'CE-08', 'CE-09',
            'CS-01', 'CS-02', 'CS-03', 'CS-04', 'CS-05', 'CS-06',
            'EC-01', 'EC-02', 'EC-03', 'EC-04', 'EC-05', 'EC-06', 'EC-07', 'EC-08', 'EC-09',
            'EE-01', 'EE-02', 'EE-03', 'EE-04', 'EE-05', 'EE-06', 'EE-07', 'EE-08', 'EE-09',
            'IT-01', 'IT-02', 'IT-03', 'IT-04', 'IT-05', 'IT-06',
            'ME-01', 'ME-02', 'ME-03', 'ME-04', 'ME-05', 'ME-06', 'ME-07', 'ME-08', 'ME-09',
            'PI-01', 'PI-02', 'PI-03', 'PI-04', 'PI-05', 'PI-06'
        ]

intents = discord.Intents.all()
client = commands.Bot(command_prefix = prefix, intents = intents, help_command=commands.DefaultHelpCommand())
client.add_cog(Tags())

available = '\nAvailable commands: `' + prefix + 'verify`, `' + prefix + 'profile`, `' + prefix + 'memlist`, `' + prefix + 'tag` and `' + prefix + 'vf_start`.'

tag = '''
Now you're able to tag roles of subsections _given_ that the said subsection falls in the same section that you are in.
That means that if you're in IT-A, you can tag `IT-A, IT-01, IT-02, ...` but you can NOT tag `IT-B, IT-06, ME-B, CE-01, PI-06, ...`

**How do I tag? The normal tagging still doesn't work.**
Here's how, you will have to precede your message with `%tag ` and type the role tag manually along with the message that you want to send.
Examples:
`%tag Hey @CS-01 can you solve this?`
`%tag @it-a when's the due date of the assignment`
Capitalization does NOT matter.

**PS:** If you're found to abuse this facility, ie. spam tags or tag people for an unimportant reason, then this facility will be revoked for you and you may be banned temporarily. Though I'll try to avoid such things. (And you should too)'''

vf = '''
This is a recreatation of the Voltorb Flip game that appears in the Korean and Western releases of Pokémon HeartGold and SoulSilver. The game is a mix between Minesweeper and Picture Cross and the placement of the bombs are given for each row and column. The goal of the game is to uncover all of the 2 and 3 tiles on a given board and move up to higher levels which have higher coin totals.

The numbers on the side and bottom of the game board denote the sum of the tiles and how many bombs are present in that row/column, respectively. Each tile you flip multiplies your collected coins by that value. Once you uncover all of the 2 and 3 tiles, all of the coins you gained this level will be added to your total and you'll go up one level to a max of 7. If you flip over a Voltorb, you lose all your coins from the current level and risk going down to a lower level.'''

dm_message = '''Welcome to the NITKKR\'24 server! Before you can see/use all the channels that it has, you'll need to do a quick verification. The process of which is explained in the #welcome channel of the server. Please do not send the command to this dm as it will not be read, instead send it on the #commands channel on the server. If you have any issues with the command, @Priyanshu will help you out personally on the channel. But do try even if you didn't understand. Have fun!'''

ft = Font(color = '0000FF00')
ft_reset = Font(color = '00000000')

class bcolors:
    Purple = '\033[95m'
    Blue = '\033[94m'
    Green = '\033[92m'
    Yellow = '\033[93m'
    Red = '\033[91m'
    White = '\033[0m'
    Bold = '\033[1m'
    Underline = '\033[4m'

@client.event
async def on_ready():
    print(f'Logged on as {client.user}!\n')

@client.event
async def on_member_join(member):
    conn = sqlite3.connect('db/details.db')
    c = conn.cursor()
    # Checks if the user who joined is already in the database or not
    c.execute('SELECT * from main where Discord_UID = (:uid)', {'uid': member.id})
    tuple = c.fetchone()
    guild = member.guild
    if tuple:
        # Fetches the mutual guilds list from the user
        guilds = json.loads(tuple[10])
        # Adds the new guild id if it's a new one
        if guild.id not in guilds:
            guilds.append(guild.id)
        guilds = json.dumps(guilds)
        # Assigning one SubSection and one Section role to the user
        role = get(guild.roles, name = tuple[2])
        await member.add_roles(role)
        role = get(guild.roles, name = tuple[3])
        await member.add_roles(role)
        # Updating the record in the database
        c.execute('UPDATE main SET Guilds = (:guilds) where Discord_UID = (:uid)', {'uid': member.id, 'guilds': guilds})
        conn.commit()
        return
    # Adding the 'Not-Verified' role if the user details do not exist in the database
    role = get(guild.roles, name = 'Not-Verified')
    await member.add_roles(role)
    # Sends a dm to the new user explaining that they have to verify
    await member.send(dm_message)

@client.command(help='Verifies your presence in the record and gives you roles based on your section/subsection.\nAlso enables you to use the `%profile` and `%tag` commands')
async def verify(ctx):
    # Checks if the author is already in the database
    conn = sqlite3.connect('db/details.db')
    temp = conn.cursor()
    temp.execute('SELECT * from main where Discord_UID = (:uid)', {'uid': ctx.author.id})
    if temp.fetchone():
        await ctx.send('You\'re already verified!')
        return
    # Assigns message content to variable
    try:
        content = ctx.message.content.split('verify ')[1]
    except:
        await ctx.send(f'Type something after `{prefix}verify`, {ctx.author.mention}')
        return
    try:
        section = content.split(' ')[0]
        roll_no = int(content.split(' ')[1])
        c = conn.cursor()
        # Gets the record of the given roll number
        c.execute('SELECT * from main where Roll_Number = (:roll)', {'roll': roll_no})
        tuple = c.fetchone()
        # Exit if roll number doesn't exist
        if not tuple:
            await ctx.send(f'The requested record was not found, {ctx.author.mention}. Please re-check the entered details and try again')
            return
        # Exit if entered section doesn't match an existing section
        if section not in sections:
            await ctx.send(f'"{section}" is not an existing section, {ctx.author.mention}.\nPlease re-check the entered details and try again')
            return
        # Exit if entered section doesn't match the section that the roll number is bound to
        if section != tuple[2]:
            await ctx.send(f'The section that you entered does not match that of the roll number that you entered, {ctx.author.mention}.\nPlease re-check the entered details and try again')
            return
        # Exit if the record is already claimed by another user
        if tuple[9]:
            await ctx.send(f'The details you entered is of a record already claimed by `{tuple[9]}` {ctx.author.mention}.\nTry another record. If you think this was a mistake, contact a moderator.')
            return
        # Assigning one SubSection and one Section role to the user
        role = get(ctx.guild.roles, name = tuple[2])
        await ctx.author.add_roles(role)
        role = get(ctx.guild.roles, name = tuple[3])
        await ctx.author.add_roles(role)
        await ctx.send(f'Your record was found and verified {ctx.author.mention}!\nYou will now be removed from this channel.')
        # Removing the 'Not-Verified' role from the user
        role = get(ctx.guild.roles, name = 'Not-Verified')
        await ctx.author.remove_roles(role)
        # Updating the record in the database
        c.execute('UPDATE main SET Discord_UID = (:uid) WHERE Roll_Number = (:roll)', {'uid': ctx.author.id, 'roll': roll_no})
        conn.commit()
        # Changing the nick of the user to their first name
        word = tuple[4].split(' ')[0]
        await ctx.author.edit(nick = word[:1] + word[1:].lower())
    except Exception as error:
        print(f'{bcolors.Red}{error}{bcolors.White}\n')
        await ctx.send(f'Error while matching details in record {ctx.author.mention}.\nYou\'ve either entered details that don\'t match or the syntax of the command is incorrect!')

@client.command(help='Displays details of the user related to the server and the college', aliases=['p', 'prof'])
async def profile(ctx):
    try:
        member = ctx.message.mentions[0]
        if member == ctx.author:
            pass
        elif 'mod' not in [name.name for name in ctx.author.roles]:
            await ctx.send('Lol nice try (You can\'t see other\'s profiles)')
            return
    except:
        member = ctx.author
    id = str(member)
    section = str()
    for role in member.roles:
        if str(role.color) == '#f1c40f':
            section = role.name
    if not section:
        await ctx.send('The requested record wasn\'t found!')
    else:
        wb = openpyxl.load_workbook('db/' + str(ctx.guild)  + ' ' + str(ctx.guild.id) + '.xlsx')
        ws = wb[section]
        for i in range(3, 90):
            if id == ws['F' + str(i)].value:
                flag = i
                break
        roles = ', '.join([role.mention for role in member.roles if section != role.name and ws['C' + str(i)].value != role.name and '@everyone' != role.name])
        if not roles:
            roles = 'None taken'
        embed = discord.Embed(
            title = ' '.join([word[:1] + word[1:].lower() for word in ws['D' + str(i)].value.split(' ')]),
            description = '**Roll Number: **' + str(ws['B' + str(i)].value)
            + '\n**Section: **' + section + ws['C' + str(i)].value[4:]
            + '\n**Roles: **' + roles
            + '\n**Email: **' + ws['E' + str(i)].value,
            colour = member.top_role.color
        )
        embed.set_author(name = id + '\'s Profile', icon_url = member.avatar_url)
        embed.set_thumbnail(url = member.avatar_url)
        embed.set_footer(text = 'Joined at: ' + str(member.joined_at)[8:10] + '-' + str(member.joined_at)[5:7] + '-' + str(member.joined_at)[:4])
        await ctx.send(embed = embed)

@client.command(help='Displays the total number of joined and remaining students for each section. Also displays the number of losers who didn\'t verify and total number of members on this server')
async def memlist(ctx):
    list = ['CE-A', 'CE-B', 'CE-C', 'CS-A', 'CS-B', 'EC-A', 'EC-B', 'EC-C', 'EE-A', 'EE-B', 'EE-C', 'IT-A', 'IT-B', 'ME-A', 'ME-B', 'ME-C', 'PI-A', 'PI-B', 'Not-Verified']
    total = [64, 61, 64, 58, 61, 59, 59, 56, 60, 57, 55, 64, 64, 67, 69, 70, 58, 59, 1105 - len([member for member in ctx.guild.members if discord.utils.get(ctx.guild.roles, name = 'Not-Verified') not in member.roles and not member.bot])]
    previous = list[0][:2]
    no = '```lisp\n'
    for section, num in zip(list[:-1], total[:-1]):
        if previous != section[:2]:
            no += '\n'
        no += '(' + section + ' --> ' + str(str(len(discord.utils.get(ctx.guild.roles, name = section).members)) + ' joined : ' + str(num - len(discord.utils.get(ctx.guild.roles, name = section).members)) + ' remaining') + ')\n'
        previous = section[:2]
    no += '\n(Verified --> ' + str(len([member for member in ctx.guild.members if discord.utils.get(ctx.guild.roles, name = 'Not-Verified') not in member.roles and not member.bot])) + ')\n'
    no += '\n'.join(['(' + role + ' --> ' + str(str(len(discord.utils.get(ctx.guild.roles, name = role).members)) + ' joined : ' + str(i - len(discord.utils.get(ctx.guild.roles, name = role).members)) + ' remaining') + ')' for role, i in zip(list[18:], total[18:])]) + '\n'
    no += '(Total --> ' + str(len([member for member in ctx.guild.members if not member.bot])) + ')```'
    await ctx.send(no)

@client.command(help='Use this to tag the subsection roles of your section.\n\n**How to use:**\n' + tag)
async def tag(ctx):
    try:
        msg = ctx.message.content.split('tag ')[1]
    except:
        await ctx.send('Type something after `' + prefix + 'tag`')
        return
    bool = False
    async with aiohttp.ClientSession() as session:
        webhooks = await ctx.channel.webhooks()
        for webhook in webhooks:
            if webhook.user == client.user:
                bool = True
                break
        if not bool:
            webhook = await ctx.channel.create_webhook(name='Webhook')
            print('created hook')
    if ctx.author.nick:
        user = ctx.author.nick
    else:
        user = str(ctx.author.name)
    flag = True
    for role in ctx.author.roles:
        if str(role.color) == '#f1c40f':
            section = role.name
    for i in msg.split(' '):
        if '@everyone' in i or '@here' in i:
            await ctx.send('You can\'t tag `@everyone` or `@here`')
            return
        if i and '@' in i:
            i = i.replace('\\', '')
            for j in i.split('@')[1:]:
                usertag = False
                for k in ctx.message.mentions:
                    if str(k.id) in j:
                        usertag = True
                if '&' in j and int(j[1:-1]) in [role.id for role in ctx.guild.roles]:
                    msg = msg.replace('\<@' + j, '@' + ctx.guild.get_role(int(j[1:-1])).name)
                    j = j.replace(j, ctx.guild.get_role(int(j[1:-1])).name)
                elif j[:4].upper() not in sections and j[:5].upper() not in subsections:
                    break
                if not usertag:
                    if j and j[:2].upper() in section:
                        if j[3] == '0':
                            if section[3] == 'A' and (j[4] == '1' or j[4] == '2' or j[4] == '3'):
                                msg = msg.replace('@' + j[:5], discord.utils.get(ctx.guild.roles, name = j[:5].strip().upper()).mention)
                            elif section[3] == 'B' and (j[4] == '4' or j[4] == '5' or j[4] == '6'):
                                msg = msg.replace('@' + j[:5], discord.utils.get(ctx.guild.roles, name = j[:5].strip().upper()).mention)
                            elif section[3] == 'C' and (j[4] == '7' or j[4] == '8' or j[4] == '9'):
                                msg = msg.replace('@' + j[:5], discord.utils.get(ctx.guild.roles, name = j[:5].strip().upper()).mention)
                            else:
                                print('1')
                                await ctx.send('You can\'t tag sections other than your own!')
                                return
                        elif j[3].upper() == section[3]:
                            msg = msg.replace('@' + j[:4], discord.utils.get(ctx.guild.roles, name = j[:4].strip().upper()).mention)
                        else:
                            print('2')
                            await ctx.send('You can\'t tag sections other than your own!')
                            return
                    elif j:
                        print('3')
                        await ctx.send('You can\'t tag sections other than your own!')
                        return

    if flag:
        await ctx.message.delete()
        await webhook.send(msg.strip(), username=user, avatar_url=ctx.author.avatar_url)

class Voltorb:
    def __init__(self):
        self.level = 1
        self.coins = 0
        self.total = 0
        self.lose = False
        self.win = False
        self.guild = discord.Guild
        self.channel = discord.TextChannel
        self.rip = discord.Message
        self.member = discord.Member
        self.message = discord.Message
        self.spam = client.get_channel(810764987076968500)
        self.url = ''
        self.vol = voltorb()
        self.row = ['a', 'b', 'c', 'd', 'e']
        self.col = [1, 2, 3, 4, 5]

    async def run(self, ctx):
        self.channel = ctx.channel
        self.guild = ctx.guild
        self.member = ctx.author
        thumb = discord.File('voltorb.gif', filename='voltorb.gif')
        board = self.vol.create(name=str(self.member))
        with BytesIO() as image_binary:
            board.save(image_binary, "PNG")
            image_binary.seek(0)
            board = discord.File(fp = image_binary, filename='board.png')
        embed = discord.Embed(
            title = 'Voltorb Flip',
            colour = self.member.top_role.color
        )
        embed.set_author(name = str(self.member) + '\'s session', icon_url = self.member.avatar_url)
        embed.set_thumbnail(url = 'attachment://voltorb.gif')
        embed.set_image(url = 'attachment://board.png')
        embed.add_field(name='Level:', value=str(self.level), inline=True)
        embed.add_field(name='Coins:', value=str(self.coins), inline=True)
        embed.add_field(name='Total Coins:', value=str(self.total), inline=True)
        await ctx.message.delete()
        self.message = await self.channel.send(files=[thumb, board], embed = embed)

    async def flip(self, ctx):
        try:
            key = ctx.message.content.split('flip ')[1]
        except IndexError:
            await ctx.send('Type something after the command')
            return
        if 'row' in key[:3].lower() or 'col' in key[:3].lower():
            coins, board = self.vol.edit_all(str(self.member), key[:3].lower(), key[4:].lower())
        else:
            if key[0] not in self.row or int(key[1]) not in self.col:
                await ctx.message.delete()
                bruh = await ctx.send('That\'s an invalid tile')
                await bruh.delete(delay=5)
                return
            coins, board = self.vol.edit(str(self.member), key)
        try:
            int(coins)
            if coins == -1:
                self.lose = True
            elif self.coins == 0:
                self.coins = coins
            elif not coins:
                await ctx.send('The tile(s) is/are already flipped!')
                return
            else:
                self.coins *= coins
        except ValueError:
            self.win = True
            self.total += self.coins*int(coins[:-1])
        thumb = discord.File('voltorb.gif', filename='voltorb.gif')
        with BytesIO() as image_binary:
            board.save(image_binary, "PNG")
            image_binary.seek(0)
            board = discord.File(fp = image_binary, filename='board.png')
        embed = discord.Embed(
            title = 'Voltorb Flip',
            colour = self.member.top_role.color
        )
        embed.set_author(name = str(self.member) + '\'s session', icon_url = self.member.avatar_url)
        embed.set_thumbnail(url = 'attachment://voltorb.gif')
        embed.add_field(name='Level:', value=str(self.level), inline=True)
        embed.add_field(name='Coins:', value=str(self.coins), inline=True)
        embed.add_field(name='Total Coins:', value=str(self.total), inline=True)
        embed.set_image(url = 'attachment://board.png')
        await ctx.message.delete()
        await self.message.delete()
        self.message = await self.channel.send(files=[thumb, board], embed = embed)
        if self.lose:
            self.rip = await ctx.send('Oh no! You hit a voltorb, ' + ctx.author.mention + ' and got 0 coins!\nType `%resume` to continue.')
        if self.win:
            self.rip = await ctx.send('Game clear, ' + ctx.author.mention + '! You received ' + str(self.coins*int(coins[:-1])) + ' Coins! Type `' + prefix + 'advance` to advance to level ' + str(self.level + 1))
            self.coins = 0

    async def resume(self, ctx):
        try:
            await self.rip.delete()
            if self.level != 1:
                self.level -= 1
            self.coins = 0
            self.vol = voltorb()
            await self.message.delete()
            await self.run(ctx)
            self.lose = False
        except TypeError:
            await ctx.send('You didn\'t lose your current match yet, ' + ctx.author.mention + '. If you would like to restart, type `' + prefix + 'restart`')

    async def advance(self, ctx):
        await self.rip.delete()
        self.level += 1
        self.vol = voltorb()
        await self.message.delete()
        await self.run(ctx)
        self.win = False

d = {}

@client.command()
async def excel(ctx):
    list = ['CE-A', 'CE-B', 'CE-C', 'CS-A', 'CS-B', 'EC-A', 'EC-B', 'EC-C', 'EE-A', 'EE-B', 'EE-C', 'IT-A', 'IT-B', 'ME-A', 'ME-B', 'ME-C', 'PI-A', 'PI-B']
    total = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    joined = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    wb = openpyxl.load_workbook('db/' + str(ctx.guild)  + ' ' + str(ctx.guild.id) + '.xlsx', read_only = True)
    for i, j in zip(list, range(len(list))):
        ws = wb[i]
        for k in range(3, 80):
            if ws['B' + str(k)].value:
                total[j] += 1
                if ws['F' + str(k)].value:
                    joined[j] += 1
                total[-1] += 1
    wb.close()
    joined[-1] = sum(joined[:-1])
    no1 = '\n'.join(['(' + list[i] + ' --> ' + str(joined[i]) + ' joined : ' + str(total[i] - joined[i]) + ' remaining)' for i in range(3)]) + '\n\n'
    no2 = '\n'.join(['(' + list[i] + ' --> ' + str(joined[i]) + ' joined : ' + str(total[i] - joined[i]) + ' remaining)' for i in range(3, 5)]) + '\n\n'
    no3 = '\n'.join(['(' + list[i] + ' --> ' + str(joined[i]) + ' joined : ' + str(total[i] - joined[i]) + ' remaining)' for i in range(5, 8)]) + '\n\n'
    no4 = '\n'.join(['(' + list[i] + ' --> ' + str(joined[i]) + ' joined : ' + str(total[i] - joined[i]) + ' remaining)' for i in range(8, 11)]) + '\n\n'
    no5 = '\n'.join(['(' + list[i] + ' --> ' + str(joined[i]) + ' joined : ' + str(total[i] - joined[i]) + ' remaining)' for i in range(11, 13)]) + '\n\n'
    no6 = '\n'.join(['(' + list[i] + ' --> ' + str(joined[i]) + ' joined : ' + str(total[i] - joined[i]) + ' remaining)' for i in range(13, 16)]) + '\n\n'
    no7 = '\n'.join(['(' + list[i] + ' --> ' + str(joined[i]) + ' joined : ' + str(total[i] - joined[i]) + ' remaining)' for i in range(16, 18)]) + '\n'
    no8 = '(Not-Verified --> ' + str(joined[-1]) + ' joined : ' + str(total[-1] - joined[-1]) + ' remaining)\n'
    await ctx.send('```lisp\n' + no1 + no2 + no3 + no4 + no5 + no6 + no7 + '\n(Verified --> ' + str(len([member for member in ctx.guild.members if discord.utils.get(ctx.guild.roles, name = 'Not-Verified') not in member.roles and not member.bot])) + ')\n' + no8 + '(Total --> ' + str(len([member for member in ctx.guild.members if not member.bot])) + ')```')

@client.event
async def on_command_error(ctx, error):
    if 'Command' in str(error) and 'is not found' in str(error):
        await ctx.send(str(error) + available)
    else:
        print(f'{bcolors.Red}{error}{bcolors.White}\n')
        await ctx.send('\nAn error occurred, contact ' + client.get_guild(783215699707166760).get_member(534651911903772674).mention + available)
        raise error

@client.command()
async def rename(ctx):
    if 'mod' not in [name.name for name in ctx.author.roles] or not ctx.message.mentions:
        await ctx.send('Lol nice try')
        return
    member = ctx.message.mentions[0]
    id = str(member)
    section = str()
    for role in member.roles:
        if str(role.color) == '#f1c40f':
            section = role.name
    if not section:
        await ctx.send('The requested record wasn\'t found!')
    else:
        wb = openpyxl.load_workbook('db/' + str(ctx.guild)  + ' ' + str(ctx.guild.id) + '.xlsx')
        ws = wb[section]
        for i in range(3, 90):
            if id == ws['F' + str(i)].value:
                word = ws['D' + str(i)].value.split(' ')[0]
                await member.edit(nick = word[:1] + word[1:].lower())
                break
    await ctx.message.delete()

@client.event
async def on_member_remove(member):
    if not member.bot:
        try:
            for role in member.roles:
                if str(role.color) == '#f1c40f':
                    section = role.name
                    break
            print(section)
            wb = openpyxl.load_workbook('db/' + str(member.guild)  + ' ' + str(member.guild.id) + '.xlsx')
            ws = wb[section]
            for i in range(3, 90):
                print(member, ws['F' + str(i)].value)
                if str(member) == ws['F' + str(i)].value:
                    ws['B' + str(i)].font = ft_reset
                    ws['C' + str(i)].font = ft_reset
                    ws['D' + str(i)].font = ft_reset
                    ws['E' + str(i)].font = ft_reset
                    ws['F' + str(i)].font = ft_reset
                    ws['F' + str(i)] = ''
                    break
            channel = client.get_channel(783215699707166763)
            await channel.send(f'**{member}** has left the server. I guess they just didn\'t like it ¯\_(ツ)_/¯')
            wb.save('db/' + str(member.guild)  + ' ' + str(member.guild.id) + '.xlsx')
        except Exception as error:
            channel = client.get_channel(783215699707166763)
            await channel.send(f'**{member}** has left the server without even verifying <a:triggered:803206114623619092>')
            raise error

@client.event
async def on_user_update(old, new):
    print(old.name)
    old = old.mutual_guilds[0].get_member(old.id)
    if old.name == new.name and old.id == new.id:
        return
    section = str()
    for role in old.roles:
        if str(role.color) == '#f1c40f':
            section = role.name
    wb = openpyxl.load_workbook('db/' + str(old.guild)  + ' ' + str(old.guild.id) + '.xlsx')
    ws = wb[section]
    for i in range(3, 90):
        if id == ws['F' + str(i)].value:
            print('\nChanged the ID in database\n')
            ws['F' + str(i)] = str(new)
    wb.save('db/' + str(old.guild)  + ' ' + str(old.guild.id) + '.xlsx')

@client.command(help=vf, aliases=['vf_start'])
async def voltorb_start(ctx):
    v1 = Voltorb()
    d[ctx.author.id] = v1
    await d[ctx.author.id].run(ctx=ctx)

@client.command()
async def flip(ctx):
    try:
        if d[ctx.author.id].win:
            await ctx.send('You\'ve already won your current session ' + ctx.author.mention + '. Type `' + prefix + 'advance` to proceed to the next level')
            return
        if d[ctx.author.id].lose:
            await ctx.send('You\'ve lost your current session ' + ctx.author.mention + '. Type `' + prefix + 'resume` to continue')
            return
        await d[ctx.author.id].flip(ctx=ctx)
    except KeyError:
        await ctx.send('You didn\'t start playing, ' + ctx.author.mention + '. Type `%vf_start` to get started.')

@client.command()
async def resume(ctx):
    try:
        if d[ctx.author.id].win:
            await ctx.send('You\'ve already won your current session ' + ctx.author.mention + '. Type `' + prefix + 'advance` to proceed to the next level')
            return
        if not d[ctx.author.id].lose:
            await ctx.send('You\'ve not lost your current session yet ' + ctx.author.mention + '.')
            return
        await d[ctx.author.id].resume(ctx=ctx)
    except KeyError:
        await ctx.send('You didn\'t start playing, ' + ctx.author.mention + '. Type `' + prefix + 'vf_start` to get started.')

@client.command()
async def advance(ctx):
    try:
        if d[ctx.author.id].win:
            await d[ctx.author.id].advance(ctx=ctx)
            d[ctx.author.id].win = False
        elif d[ctx.author.id].lose:
            await ctx.send('You\'ve lost your current session ' + ctx.author.mention + '. Type `' + prefix + 'resume` to continue')
            return
        else:
            await ctx.send('You didn\'t win your current match yet, ' + ctx.author.mention + '. If you would like to restart, type `' + prefix + 'restart`')
    except KeyError:
        await ctx.send('You didn\'t start playing, ' + ctx.author.mention + '. Type `' + prefix + 'vf_start` to get started.')

@client.command()
async def quit(ctx):
    try:
        await d[ctx.author.id].message.delete()
        d.pop(ctx.author.id)
        await ctx.message.delete()
    except KeyError:
        await ctx.send('You didn\'t start playing, ' + ctx.author.mention + '. Type `' + prefix + 'vf_start` to get started.')

@client.command(aliases=['inv'])
async def invite(ctx):
    await ctx.send('**NITKKR server:** https://discord.gg/4eF7R6afqv\n**kkr++ server:** https://discord.gg/epaTW7tjYR')

class MyHelp(commands.MinimalHelpCommand):
    async def send_command_help(self, command):
        embed = discord.Embed(title=self.get_command_signature(command))
        embed.add_field(name="Help", value=command.help)
        alias = command.aliases
        if alias:
            embed.add_field(name="Aliases", value=", ".join(alias), inline=False)

        channel = self.get_destination()
        await channel.send(embed=embed)

client.help_command = MyHelp()

client.run(os.getenv("BOT_TOKEN"))
